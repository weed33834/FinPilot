"""生成 E2E 测试用的复杂业务数据。

覆盖：
1. 腾讯控股 2025 年 Q3 财报 Excel（3 个 sheet：利润表 / 资产负债表 / 现金流量表）
2. 阿里巴巴 2025 财报 PDF（含正文 + 财务数据表）
3. 行业对标数据 Excel（5 家互联网公司横向对比）
4. 极限 case：试算不平衡 Excel / 时间穿越凭证 PDF
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# 1. 腾讯控股 2025 Q3 财报 Excel
# ---------------------------------------------------------------------------

def gen_tencent_excel() -> Path:
    """腾讯 2025 Q3 财报 Excel：3 sheet（利润表/资产负债表/现金流量表）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(name="微软雅黑", size=14, bold=True, color="1F4E78")
    border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )

    def style_sheet(ws, title: str, headers: list[str], rows: list[list]) -> None:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value=title).font = title_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
            c.border = border
        for r_idx, row in enumerate(rows, start=3):
            for c_idx, val in enumerate(row, start=1):
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                c.border = border
                c.alignment = Alignment(horizontal="left" if c_idx == 1 else "right")
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

    # Sheet 1: 利润表（单位：百万元）
    ws1 = wb.active
    ws1.title = "利润表"
    style_sheet(
        ws1,
        "腾讯控股 2025 年 Q3 利润表（单位：百万元 RMB）",
        ["项目", "2025-Q3", "2025-Q2", "2025-Q1", "2024-Q3", "同比变化"],
        [
            ["营业收入", 167211, 161122, 159501, 154625, "+8.13%"],
            ["  增值服务", 84219, 81735, 80881, 77912, "+8.09%"],
            ["    国内游戏", 42108, 40876, 40522, 38289, "+10.02%"],
            ["    国际游戏", 18394, 17821, 17530, 16233, "+13.32%"],
            ["    社交网络", 23717, 23038, 22829, 23390, "+1.40%"],
            ["  网络广告", 32467, 31852, 31685, 30045, "+8.06%"],
            ["  金融科技与企业服务", 54180, 52102, 51045, 49543, "+9.36%"],
            ["营业成本", -84217, -81933, -81060, -78219, "+7.67%"],
            ["毛利润", 82994, 79189, 78441, 76406, "+8.62%"],
            ["销售费用", -9521, -9218, -9087, -8670, "+9.81%"],
            ["管理费用", -12338, -12098, -11942, -11223, "+9.94%"],
            ["研发费用", -18654, -18321, -18102, -17265, "+8.05%"],
            ["营业利润", 42481, 39552, 39310, 39248, "+8.24%"],
            ["财务收入", 4128, 3987, 3892, 3210, "+28.60%"],
            ["财务费用", -2187, -2098, -2031, -1956, "+11.81%"],
            ["利润总额", 44422, 41441, 41171, 40502, "+9.68%"],
            ["所得税", -8201, -7632, -7567, -7456, "+10.00%"],
            ["净利润", 36221, 33809, 33604, 33046, "+9.62%"],
            ["  归母净利润", 35898, 33520, 33308, 32763, "+9.57%"],
            ["  少数股东损益", 323, 289, 296, 283, "+14.13%"],
            ["基本EPS（元）", 3.83, 3.58, 3.56, 3.50, "+9.43%"],
        ],
    )

    # Sheet 2: 资产负债表
    ws2 = wb.create_sheet("资产负债表")
    style_sheet(
        ws2,
        "腾讯控股 2025 年 Q3 资产负债表（单位：百万元 RMB）",
        ["项目", "2025-Q3", "2025-Q2", "2025-Q1", "2024-Q4", "环比变化"],
        [
            ["货币资金", 285421, 276530, 268904, 254102, "+3.22%"],
            ["应收账款", 38219, 36521, 35102, 34098, "+4.65%"],
            ["存货", 3218, 3102, 2987, 2856, "+3.74%"],
            ["其他流动资产", 65432, 62109, 59876, 57321, "+5.37%"],
            ["流动资产合计", 392290, 378262, 366869, 348377, "+3.71%"],
            ["固定资产", 45621, 44102, 42987, 41508, "+3.45%"],
            ["无形资产", 32109, 31508, 30876, 30102, "+1.91%"],
            ["商誉", 122987, 122098, 121543, 120876, "+0.73%"],
            ["长期股权投资", 287654, 282109, 276543, 269876, "+1.97%"],
            ["非流动资产合计", 712087, 698432, 687654, 673210, "+1.95%"],
            ["资产总计", 1104377, 1076694, 1054523, 1021587, "+2.57%"],
            ["短期借款", 18209, 17908, 17543, 17209, "+1.68%"],
            ["应付账款", 42318, 41098, 40102, 38987, "+2.97%"],
            ["其他流动负债", 56521, 55102, 53987, 52409, "+2.57%"],
            ["流动负债合计", 117048, 114108, 111632, 108605, "+2.58%"],
            ["长期借款", 65432, 64219, 63102, 62098, "+1.74%"],
            ["应付债券", 32109, 31876, 31543, 31209, "+0.73%"],
            ["其他非流动负债", 21876, 21309, 20876, 20432, "+2.66%"],
            ["非流动负债合计", 119417, 117404, 115521, 113739, "+1.72%"],
            ["负债合计", 236465, 231512, 227153, 222344, "+2.14%"],
            ["股本", 9382, 9382, 9382, 9382, "0.00%"],
            ["资本公积", 312987, 309876, 307654, 304321, "+1.00%"],
            ["盈余公积", 98765, 96543, 94321, 92109, "+2.30%"],
            ["未分配利润", 445778, 429381, 416013, 395431, "+3.82%"],
            ["归母所有者权益合计", 866912, 845182, 827170, 801243, "+2.57%"],
            ["少数股东权益", 1000, 999, 999, 999, "+0.10%"],
            ["所有者权益合计", 867912, 846181, 828169, 802242, "+2.57%"],
            ["负债和所有者权益总计", 1104377, 1076693, 1055322, 1024586, "+2.57%"],
        ],
    )

    # Sheet 3: 现金流量表
    ws3 = wb.create_sheet("现金流量表")
    style_sheet(
        ws3,
        "腾讯控股 2025 年 Q3 现金流量表（单位：百万元 RMB）",
        ["项目", "2025-Q3", "2025-Q2", "2025-Q1", "2024-Q3", "同比变化"],
        [
            ["销售商品提供劳务收到的现金", 184921, 178543, 176321, 168907, "+9.48%"],
            ["收到的税费返还", 1987, 1932, 1902, 1823, "+9.00%"],
            ["收到其他与经营活动有关的现金", 5432, 5219, 5098, 4876, "+11.40%"],
            ["经营活动现金流入小计", 192340, 185694, 183321, 175606, "+9.54%"],
            ["购买商品接受劳务支付的现金", -89543, -87231, -85987, -83109, "+7.74%"],
            ["支付给职工的现金", -28765, -27987, -27543, -26321, "+9.28%"],
            ["支付的各项税费", -21543, -20987, -20654, -19876, "+8.39%"],
            ["支付其他与经营活动有关的现金", -12321, -11987, -11765, -11234, "+9.67%"],
            ["经营活动现金流出小计", -152172, -148192, -145949, -140540, "+8.28%"],
            ["经营活动产生的现金流量净额", 40168, 37502, 37372, 35066, "+14.55%"],
            ["收回投资收到的现金", 21543, 20987, 20654, 19876, "+8.39%"],
            ["取得投资收益收到的现金", 5432, 5219, 5098, 4876, "+11.40%"],
            ["投资活动现金流入小计", 26975, 26206, 25752, 24752, "+9.00%"],
            ["购建固定资产支付的现金", -8765, -8543, -8421, -8109, "+8.10%"],
            ["投资支付的现金", -32109, -31098, -30432, -29321, "+9.51%"],
            ["投资活动现金流出小计", -40874, -39641, -38853, -37430, "+9.18%"],
            ["投资活动产生的现金流量净额", -13899, -13435, -13101, -12678, "+9.63%"],
            ["吸收投资收到的现金", 1098, 1065, 1043, 1009, "+8.82%"],
            ["取得借款收到的现金", 6543, 6321, 6209, 5987, "+9.29%"],
            ["筹资活动现金流入小计", 7641, 7386, 7252, 6996, "+9.22%"],
            ["偿还债务支付的现金", -5432, -5219, -5098, -4876, "+11.40%"],
            ["分配股利支付的现金", -21876, -20987, -20654, -19876, "+10.06%"],
            ["筹资活动现金流出小计", -27308, -26206, -25752, -24752, "+10.36%"],
            ["筹资活动产生的现金流量净额", -19667, -18820, -18500, -17756, "+10.78%"],
            ["现金及现金等价物净增加额", 6602, 5247, 5771, 4632, "+42.55%"],
            ["期初现金及现金等价物余额", 278819, 284066, 278295, 280970, "-0.91%"],
            ["期末现金及现金等价物余额", 285421, 289313, 284066, 285602, "-0.06%"],
        ],
    )

    out = DATA_DIR / "腾讯控股_2025Q3财报.xlsx"
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# 2. 阿里巴巴 2025 财年报告 PDF
# ---------------------------------------------------------------------------

def gen_alibaba_pdf() -> Path:
    """阿里巴巴 2025 财年报告 PDF：含正文 + 财务数据表 + 行业分析。"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, Heading1,
        )
    except ImportError:
        # reportlab 未安装，回退到生成纯文本 PDF
        return _gen_alibaba_pdf_simple()

    out = DATA_DIR / "阿里巴巴_2025财年报告.pdf"
    doc = SimpleDocTemplate(
        str(out), pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="ChineseTitle", parent=styles["Title"],
        fontName="Helvetica-Bold", fontSize=20, spaceAfter=14,
    ))
    styles.add(ParagraphStyle(
        name="ChineseBody", parent=styles["BodyText"],
        fontSize=10, leading=15, spaceAfter=8,
    ))

    elements = []

    elements.append(Paragraph("Alibaba Group Holding Limited", styles["ChineseTitle"]))
    elements.append(Paragraph("2025 财年年度报告（截至 2025 年 3 月 31 日）", styles["Heading2"]))
    elements.append(Spacer(1, 0.5 * cm))

    elements.append(Paragraph("一、管理层讨论与分析", styles["Heading1"]))
    elements.append(Paragraph(
        "2025 财年，阿里巴巴集团实现总收入 9,943.11 亿元人民币，同比增长 5.9%。"
        "剔除集团内部交易后，对外收入 9,416.83 亿元，同比增长 6.4%。"
        "淘天集团客户管理收入（CMR）2,891.32 亿元，同比增长 4.7%；"
        "云智能集团收入 1,131.21 亿元，同比增长 18.7%，主要由 AI 相关产品收入 triple 增长驱动；"
        "国际数字商业集团（AIDC）收入 1,325.49 亿元，同比增长 22.5%。",
        styles["ChineseBody"],
    ))
    elements.append(Paragraph(
        "经营利润 1,325.66 亿元，同比增长 32.1%；归属于普通股股东的净利润 1,414.39 亿元，"
        "同比增长 76.5%。Non-GAAP 净利润 1,578.41 亿元，同比增长 1.4%。",
        styles["ChineseBody"],
    ))
    elements.append(Spacer(1, 0.3 * cm))

    elements.append(Paragraph("二、各业务集团收入构成", styles["Heading1"]))
    revenue_data = [
        ["业务集团", "2025 财年（百万元）", "2024 财年（百万元）", "同比变化"],
        ["淘天集团", "471,012", "447,020", "+5.4%"],
        ["云智能集团", "113,121", "105,287", "+7.4%"],
        ["国际数字商业集团", "132,549", "108,206", "+22.5%"],
        ["本地生活集团", "67,002", "59,803", "+12.0%"],
        ["菜鸟集团", "99,388", "98,047", "+1.4%"],
        ["大文娱集团", "22,321", "20,742", "+7.6%"],
        ["其他业务", "89,238", "78,978", "+13.0%"],
        ["合计", "994,311", "918,083", "+8.3%"],
    ]
    t = Table(revenue_data, colWidths=[4 * cm, 4 * cm, 4 * cm, 3 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.5 * cm))

    elements.append(Paragraph("三、关键风险因素", styles["Heading1"]))
    risk_text = (
        "1. 竞争加剧：电商市场面临拼多多、抖音电商等竞争对手压力，淘天集团市场份额被持续侵蚀。<br/>"
        "2. 监管风险：反垄断、数据安全、跨境数据传输等监管政策对公司业务产生影响。<br/>"
        "3. AI 投入产出不确定性：云智能集团 AI 基础设施投入巨大，但商业化节奏存在不确定性。<br/>"
        "4. 国际化风险：AIDC 在东南亚、欧洲面临当地合规与地缘政治风险。<br/>"
        "5. 关联交易披露：蚂蚁集团回购股份后，与公司的关联交易需持续披露。"
    )
    elements.append(Paragraph(risk_text, styles["ChineseBody"]))
    elements.append(Spacer(1, 0.3 * cm))

    elements.append(Paragraph("四、未来发展战略", styles["Heading1"]))
    elements.append(Paragraph(
        "1. <b>淘天集团</b>：聚焦用户体验，加大 88VIP 用户运营投入，GMV 同比增长目标 +8%。<br/>"
        "2. <b>云智能</b>：以 AI 为核心驱动力，未来 3 年 AI 相关产品收入年复合增长目标 +50%。<br/>"
        "3. <b>AIDC</b>：聚焦东南亚六国 + 欧洲重点市场，AliExpress 与 Lazada 双品牌协同。<br/>"
        "4. <b>本地生活</b>：饿了么 UE 持续改善，目标 2026 财年实现整体盈亏平衡。<br/>"
        "5. <b>菜鸟</b>：加速全球物流网络建设，跨境电商履约时效提升 20%。",
        styles["ChineseBody"],
    ))

    elements.append(PageBreak())
    elements.append(Paragraph("五、财务数据附录", styles["Heading1"]))

    appendix_data = [
        ["指标", "2025财年", "2024财年", "2023财年"],
        ["总收入（百万元）", "994,311", "918,083", "868,687"],
        ["经营利润（百万元）", "132,566", "100,351", "68,642"],
        ["净利润（百万元）", "141,439", "80,099", "65,573"],
        ["毛利率", "41.8%", "40.5%", "39.2%"],
        ["营业利润率", "13.3%", "10.9%", "7.9%"],
        ["净利率", "14.2%", "8.7%", "7.6%"],
        ["经营活动现金流（百万元）", "178,902", "152,876", "138,543"],
        ["资本开支（百万元）", "-85,432", "-62,109", "-48,765"],
        ["自由现金流（百万元）", "93,470", "90,767", "89,778"],
        ["ROE", "12.8%", "8.4%", "7.5%"],
        ["资产负债率", "37.5%", "38.2%", "39.8%"],
    ]
    t2 = Table(appendix_data, colWidths=[5 * cm, 3.5 * cm, 3.5 * cm, 3.5 * cm])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elements.append(t2)

    doc.build(elements)
    return out


def _gen_alibaba_pdf_simple() -> Path:
    """fallback：用 fpdf 生成纯文本 PDF。"""
    from fpdf import FPDF

    out = DATA_DIR / "阿里巴巴_2025财年报告.pdf"
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)
    pdf.cell(200, 10, "Alibaba Group 2025 Annual Report", ln=True, align="C")
    pdf.ln(5)
    content = """Alibaba Group Holding Limited 2025 Annual Report

1. Revenue: RMB 994.31 billion (+8.3% YoY)
2. Operating profit: RMB 132.57 billion (+32.1% YoY)
3. Net profit: RMB 141.44 billion (+76.5% YoY)
4. Gross margin: 41.8%
5. Operating margin: 13.3%

Business segments:
- Taobao Tmall: 471.01 billion (+5.4%)
- Cloud Intelligence: 113.12 billion (+7.4%)
- International Digital Commerce: 132.55 billion (+22.5%)
- Local Services: 67.00 billion (+12.0%)
- Cainiao: 99.39 billion (+1.4%)
- Digital Media: 22.32 billion (+7.6%)
"""
    for line in content.split("\n"):
        pdf.cell(0, 6, line, ln=True)
    pdf.output(str(out))
    return out


# ---------------------------------------------------------------------------
# 3. 行业对标数据 Excel
# ---------------------------------------------------------------------------

def gen_industry_compare_excel() -> Path:
    """5 家互联网公司横向对比 Excel。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "行业对标"
    ws.cell(row=1, column=1, value="中国互联网 5 大厂 2025 Q3 财务对比（单位：百万元 RMB）")
    ws.cell(row=2, column=1, value="指标")
    headers = ["腾讯控股", "阿里巴巴", "字节跳动", "美团", "拼多多"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=2, column=i, value=h)

    rows = [
        ["营业收入", 167211, 247823, 254890, 89765, 102340],
        ["营收同比增速", "8.1%", "11.2%", "32.4%", "22.0%", "44.3%"],
        ["净利润", 36221, 35218, 54000, 12890, 24980],
        ["净利率", "21.7%", "14.2%", "21.2%", "14.4%", "24.4%"],
        ["毛利率", "49.6%", "41.8%", "53.2%", "37.8%", "62.1%"],
        ["经营现金流", 40168, 44520, 62109, 15876, 28765],
        ["资本开支", -8765, -21387, -32543, -5432, -2109],
        ["自由现金流", 31403, 23133, 29566, 10444, 26656],
        ["研发费用", 18654, 19321, 28987, 5432, 3210],
        ["研发费用率", "11.2%", "7.8%", "11.4%", "6.1%", "3.1%"],
        ["销售费用", 9521, 31298, 65432, 28765, 42310],
        ["销售费用率", "5.7%", "12.6%", "25.7%", "32.0%", "41.3%"],
        ["资产负债率", "21.4%", "37.5%", "32.8%", "25.6%", "18.5%"],
        ["货币资金", 285421, 421098, 504321, 187654, 198765],
        ["员工人数", 110558, 198902, 150000, 110000, 13000],
        ["人均创收（万元）", 1512, 1245, 1699, 816, 7872],
        ["人均创利（万元）", 327, 177, 360, 117, 1922],
    ]
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # 宽度
    from openpyxl.utils import get_column_letter
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

    out = DATA_DIR / "互联网5大厂_2025Q3对标.xlsx"
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# 4. 极限 case 数据
# ---------------------------------------------------------------------------

def gen_extreme_trial_balance() -> Path:
    """试算不平衡的 Excel（借贷差 1000 元，应触发 P0 校验）。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "试算不平衡"
    ws.cell(row=1, column=1, value="某公司 2025-09 月末试算平衡表（单位：元）")
    headers = ["科目代码", "科目名称", "借方", "贷方"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    rows = [
        ["1001", "库存现金", 50230.00, None],
        ["1002", "银行存款", 1234567.89, None],
        ["1122", "应收账款", 4567890.12, None],
        ["1401", "原材料", 876543.21, None],
        ["1601", "固定资产", 9876543.00, None],
        ["2001", "应付账款", None, 2345678.90],
        ["2202", "应付职工薪酬", None, 543210.00],
        ["2501", "长期借款", None, 5000000.00],
        ["4001", "实收资本", None, 8000000.00],
        ["5001", "主营业务收入", None, 765432.10],  # 故意少 1000 元 → 试算不平衡
    ]
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    # 合计行
    total_debit = sum(r[2] or 0 for r in rows)
    total_credit = sum(r[3] or 0 for r in rows)
    ws.cell(row=len(rows) + 3, column=1, value="合计")
    ws.cell(row=len(rows) + 3, column=3, value=total_debit)
    ws.cell(row=len(rows) + 3, column=4, value=total_credit)
    ws.cell(row=len(rows) + 4, column=1, value=f"差额：{total_debit - total_credit}（试算不平衡！）")

    out = DATA_DIR / "extreme_试算不平衡.xlsx"
    wb.save(out)
    return out


def gen_extreme_time_travel_pdf() -> Path:
    """时间穿越凭证 PDF：包含 2025-12-31 日期的凭证（结账日是 2025-09-30）。"""
    from fpdf import FPDF

    out = DATA_DIR / "extreme_时间穿越凭证.pdf"
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 10, "Accounting Voucher - October 2025", ln=True, align="C")
    pdf.ln(5)
    pdf.cell(0, 8, "Closing date: 2025-09-30 (FY 2025-Q3)", ln=True)
    pdf.ln(3)
    content = """Voucher No: JZ-202509-001  Date: 2025-09-15
  Debit: Raw material inventory  50000.00
  Credit: Bank deposit           50000.00

Voucher No: JZ-202509-002  Date: 2025-09-25
  Debit: Operating expense       32000.00
  Credit: Cash                   32000.00

Voucher No: JZ-202512-001  Date: 2025-12-31 (TIME TRAVEL!)
  Debit: Equipment               120000.00
  Credit: Bank deposit          120000.00

Voucher No: JZ-202508-099  Date: 2025-08-30
  Debit: Prepaid expense         15000.00
  Credit: Bank deposit           15000.00

Voucher No: JZ-202601-001  Date: 2026-01-15 (TIME TRAVEL!)
  Debit: Advance payment         80000.00
  Credit: Bank deposit           80000.00
"""
    for line in content.split("\n"):
        pdf.cell(0, 5, line, ln=True)
    pdf.output(str(out))
    return out


# ---------------------------------------------------------------------------
# 5. 上市公司财报汇总 PDF
# ---------------------------------------------------------------------------

def gen_industry_brief_pdf() -> Path:
    """行业研报 PDF：包含 5 家公司评级 + 投资建议（用于 Agent 问"给我建议"场景）。"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Heading1,
        )
    except ImportError:
        return _gen_industry_brief_pdf_simple()

    out = DATA_DIR / "互联网行业_2025Q3投资研报.pdf"
    doc = SimpleDocTemplate(
        str(out), pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="CBody", parent=styles["BodyText"], fontSize=10, leading=15, spaceAfter=6,
    ))
    elements = []
    elements.append(Paragraph("互联网行业 2025 Q3 投资策略研报", styles["Title"]))
    elements.append(Paragraph("中金公司 · 2025-11-10", styles["Heading2"]))
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph("一、核心观点", styles["Heading1"]))
    elements.append(Paragraph(
        "2025 Q3 中国互联网行业整体收入同比增长 17.4%，环比 Q2 加速 1.2 pct。"
        "AI 商业化进入兑现期，云业务成为收入增长主驱动力。"
        "电商竞争格局重塑，拼多多份额持续提升，淘天承压。"
        "建议超配<b>腾讯控股</b>、<b>字节跳动（未上市）</b>，标配<b>阿里巴巴</b>、<b>美团</b>。",
        styles["CBody"],
    ))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(Paragraph("二、个股评级与目标价", styles["Heading1"]))
    rating = [
        ["公司", "评级", "目标价(HKD)", "当前价", "上涨空间"],
        ["腾讯控股", "买入", "520", "435", "+19.5%"],
        ["阿里巴巴", "增持", "115", "98", "+17.3%"],
        ["美团", "增持", "165", "142", "+16.2%"],
        ["拼多多", "买入", "185", "152", "+21.7%"],
        ["字节跳动", "未上市", "-", "-", "-"],
    ]
    t = Table(rating, colWidths=[3.5 * cm, 2 * cm, 3 * cm, 2.5 * cm, 3 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(Paragraph("三、风险提示", styles["Heading1"]))
    elements.append(Paragraph(
        "1. 宏观经济下行风险<br/>"
        "2. 监管政策变化风险<br/>"
        "3. AI 商业化不及预期<br/>"
        "4. 国际地缘政治风险<br/>"
        "5. 关联交易披露合规风险",
        styles["CBody"],
    ))
    doc.build(elements)
    return out


def _gen_industry_brief_pdf_simple() -> Path:
    from fpdf import FPDF
    out = DATA_DIR / "互联网行业_2025Q3投资研报.pdf"
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 10, "Internet Industry 2025 Q3 Research Report", ln=True, align="C")
    pdf.ln(5)
    content = """Buy: Tencent Holdings (target 520 HKD)
Accumulate: Alibaba (target 115 HKD), Meituan (target 165 HKD)
Buy: Pinduoduo (target 185 HKD)

Key risks:
1. Macroeconomic risk
2. Regulatory risk
3. AI commercialization risk
4. Geopolitical risk
5. Related party transaction disclosure risk
"""
    for line in content.split("\n"):
        pdf.cell(0, 6, line, ln=True)
    pdf.output(str(out))
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("=== 生成 E2E 测试业务数据 ===")
    results = []
    try:
        results.append(("腾讯 Excel", gen_tencent_excel()))
    except Exception as e:
        print(f"  [FAIL] 腾讯 Excel: {e}")
    try:
        results.append(("阿里 PDF", gen_alibaba_pdf()))
    except Exception as e:
        print(f"  [FAIL] 阿里 PDF: {e}")
    try:
        results.append(("行业对标 Excel", gen_industry_compare_excel()))
    except Exception as e:
        print(f"  [FAIL] 行业对标 Excel: {e}")
    try:
        results.append(("试算不平衡 Excel", gen_extreme_trial_balance()))
    except Exception as e:
        print(f"  [FAIL] 试算不平衡 Excel: {e}")
    try:
        results.append(("时间穿越 PDF", gen_extreme_time_travel_pdf()))
    except Exception as e:
        print(f"  [FAIL] 时间穿越 PDF: {e}")
    try:
        results.append(("行业研报 PDF", gen_industry_brief_pdf()))
    except Exception as e:
        print(f"  [FAIL] 行业研报 PDF: {e}")

    print("\n=== 已生成文件 ===")
    for name, path in results:
        size_kb = path.stat().st_size / 1024
        print(f"  [OK] {name}: {path} ({size_kb:.1f} KB)")
    print(f"\n总计: {len(results)} 个文件，目录: {DATA_DIR}")


if __name__ == "__main__":
    main()
